"""
Processes MoFA SRID Excel files from the configured inbox folder.

Each Excel file may contain multiple sheets representing different
markets or regions.  Sheets flagged as cover/summary/contents pages
are skipped; all others are merged into a single DataFrame and
persisted as raw_prices rows for the validate → transform pipeline.
"""

import json
import logging
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from sqlalchemy import select

from config.settings import MOFA_INBOX_PATH
from db.connection import get_session
from db.models import IngestionLog, RawPrice

logger = logging.getLogger(__name__)

_SOURCE = "mofa_srid"
_SKIP_SHEET_KEYWORDS = {"cover", "summary", "contents"}
_BATCH = 500


class MoFAClient:

    def __init__(self) -> None:
        self.inbox: Path = MOFA_INBOX_PATH

    # ── 1. Discover unprocessed files ─────────────────────────────────────────

    def get_unprocessed_files(self) -> list[Path]:
        """Return .xlsx files in inbox that have not yet been ingested.

        A file is considered processed if its name appears as ``file_ref``
        in any successful ``ingestion_log`` row with ``source='mofa_srid'``.
        """
        if not self.inbox.exists():
            logger.warning("MoFA inbox does not exist: %s", self.inbox)
            return []

        all_files = list(self.inbox.glob("*.xlsx"))
        if not all_files:
            logger.info("No .xlsx files found in %s", self.inbox)
            return []

        logger.info("Found %d .xlsx file(s) in inbox", len(all_files))

        with get_session() as session:
            rows = session.execute(
                select(IngestionLog.file_ref).where(
                    IngestionLog.source == _SOURCE,
                    IngestionLog.status == "success",
                )
            ).scalars().all()

        processed_names = {Path(r).name for r in rows if r}
        unprocessed = [f for f in all_files if f.name not in processed_names]

        logger.info(
            "%d already processed, %d remaining",
            len(all_files) - len(unprocessed),
            len(unprocessed),
        )
        return unprocessed

    # ── 2. Locate header row ──────────────────────────────────────────────────

    def detect_header_row(self, worksheet) -> int:
        """Scan the first 10 rows for the row that contains 'market' and 'price'.

        Returns the 0-based row index suitable for ``pandas.read_excel``'s
        ``header`` parameter.  Falls back to 0 with a warning if not found.
        """
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10)):
            cell_values = [
                str(cell.value).strip().lower()
                for cell in row
                if cell.value is not None
            ]
            if any("market" in v for v in cell_values) and any(
                "price" in v for v in cell_values
            ):
                logger.debug(
                    "Header row detected at index %d in sheet '%s'",
                    row_idx,
                    worksheet.title,
                )
                return row_idx

        logger.warning(
            "Could not detect header row in sheet '%s' -- defaulting to row 0",
            worksheet.title,
        )
        return 0

    # ── 3. Parse an Excel file ────────────────────────────────────────────────

    def parse_xlsx(self, filepath: Path) -> pd.DataFrame:
        """Read all data sheets from *filepath* into a single DataFrame.

        Skips sheets whose name contains 'cover', 'summary', or 'contents'
        (case-insensitive).  Merges the remaining sheets and adds metadata
        columns ``_sheet`` and ``_source_file``.
        """
        logger.info("Opening %s", filepath.name)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        frames: list[pd.DataFrame] = []

        for sheet_name in wb.sheetnames:
            lower_name = sheet_name.strip().lower()
            if any(kw in lower_name for kw in _SKIP_SHEET_KEYWORDS):
                logger.debug("Skipping sheet '%s' (cover/summary/contents)", sheet_name)
                continue

            ws = wb[sheet_name]
            header_row = self.detect_header_row(ws)

            try:
                df = pd.read_excel(
                    filepath,
                    sheet_name=sheet_name,
                    header=header_row,
                    engine="openpyxl",
                )
            except Exception as exc:
                logger.warning(
                    "Could not read sheet '%s' in %s: %s -- skipping",
                    sheet_name,
                    filepath.name,
                    exc,
                )
                continue

            if df.empty:
                logger.debug("Sheet '%s' is empty -- skipping", sheet_name)
                continue

            # Forward-fill the first column to expand merged region/market cells.
            df.iloc[:, 0] = df.iloc[:, 0].ffill()

            df["_sheet"] = sheet_name
            df["_source_file"] = filepath.name
            frames.append(df)
            logger.info(
                "Sheet '%s': %d rows x %d columns", sheet_name, len(df), len(df.columns)
            )

        wb.close()

        if not frames:
            logger.warning("No data sheets found in %s", filepath.name)
            return pd.DataFrame()

        combined = pd.concat(frames, ignore_index=True)
        logger.info(
            "Combined %d sheet(s) -> %d rows from %s",
            len(frames),
            len(combined),
            filepath.name,
        )
        return combined

    # ── 4. Persist raw rows ───────────────────────────────────────────────────

    def save_raw_rows(self, df: pd.DataFrame, filename: str) -> list[int]:
        """Insert every DataFrame row as a RawPrice record with source='mofa_srid'.

        Returns the list of newly inserted IDs in insertion order.
        """
        records: list[dict] = json.loads(df.to_json(orient="records"))
        n = len(records)
        logger.info(
            "Inserting %d raw rows from '%s' (batch size %d)...", n, filename, _BATCH
        )

        orm_objects: list[RawPrice] = []
        with get_session() as session:
            for i, record in enumerate(records, start=1):
                obj = RawPrice(
                    source=_SOURCE,
                    raw_payload=record,
                    file_ref=filename,
                )
                session.add(obj)
                orm_objects.append(obj)

                if i % _BATCH == 0:
                    session.flush()
                    logger.debug("Flushed batch %d / %d", i, n)

            session.flush()
            raw_ids = [obj.id for obj in orm_objects]

        first, last = (raw_ids[0], raw_ids[-1]) if raw_ids else (None, None)
        logger.info("Committed %d rows (IDs %s - %s)", len(raw_ids), first, last)
        return raw_ids

    # ── 5. Orchestrate ────────────────────────────────────────────────────────

    def run(self) -> dict[str, Any]:
        """Process every unprocessed MoFA file in the inbox.

        Returns a summary dict with per-file results and an aggregate count.
        One file failing does not prevent the others from being processed.
        """
        logger.info("=== MoFA ingestion run started (inbox: %s) ===", self.inbox)

        files = self.get_unprocessed_files()
        if not files:
            logger.info("Nothing to process.")
            return {
                "source": _SOURCE,
                "files_found": 0,
                "files_processed": 0,
                "files_failed": 0,
                "total_rows_fetched": 0,
                "file_results": [],
            }

        file_results: list[dict] = []
        total_rows = 0
        files_processed = 0
        files_failed = 0

        for filepath in files:
            result: dict[str, Any] = {
                "file": filepath.name,
                "status": "failed",
                "error": None,
            }
            try:
                df = self.parse_xlsx(filepath)

                if df.empty:
                    # Record a success with 0 rows so the file is not retried.
                    self._write_log(filepath.name, rows_fetched=0, status="success")
                    result.update({"status": "success", "rows_fetched": 0, "raw_ids": []})
                    files_processed += 1
                    file_results.append(result)
                    continue

                raw_ids = self.save_raw_rows(df, filepath.name)
                rows = len(raw_ids)
                total_rows += rows

                self._write_log(filepath.name, rows_fetched=rows, status="success")
                result.update({"status": "success", "rows_fetched": rows, "raw_ids": raw_ids})
                files_processed += 1
                logger.info("File '%s' complete: %d raw rows", filepath.name, rows)

            except Exception as exc:
                logger.exception("Failed to process '%s': %s", filepath.name, exc)
                self._write_log(
                    filepath.name, rows_fetched=0, status="failed", error=str(exc)
                )
                result["error"] = str(exc)
                files_failed += 1

            file_results.append(result)

        summary: dict[str, Any] = {
            "source": _SOURCE,
            "files_found": len(files),
            "files_processed": files_processed,
            "files_failed": files_failed,
            "total_rows_fetched": total_rows,
            "file_results": file_results,
        }
        logger.info("=== MoFA ingestion complete: %s ===", summary)
        return summary

    # ── Internal ──────────────────────────────────────────────────────────────

    def _write_log(
        self,
        filename: str,
        rows_fetched: int,
        status: str,
        error: str | None = None,
    ) -> None:
        """Write one row to ingestion_log for *filename*."""
        try:
            with get_session() as session:
                session.add(
                    IngestionLog(
                        source=_SOURCE,
                        file_ref=filename,
                        rows_fetched=rows_fetched,
                        rows_clean=0,
                        rows_quarantined=0,
                        status=status,
                        error_detail=error,
                    )
                )
        except Exception as log_exc:
            logger.error(
                "Could not write ingestion_log for '%s': %s", filename, log_exc
            )
